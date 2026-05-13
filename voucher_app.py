from __future__ import annotations

import glob
import json
import logging
import os
import re
import sys
import tempfile
from datetime import datetime
from typing import Callable, Dict, Optional, cast

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from PySide6.QtCore import QObject, Qt, QThread, Signal, Slot
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QTabWidget,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

APP_TITLE = "כלי ניהול ובקרת דוקטים לגבייה"

MODERN_STYLE = """
    QMainWindow { background-color: #f5f7fa; }
    QWidget { font-family: 'Segoe UI', 'Arial'; font-size: 14px; color: #333; }
    QFrame#Card {
        background-color: white;
        border-radius: 12px;
        border: 1px solid #e1e4e8;
    }
    QLabel#HeaderLabel {
        font-size: 22px;
        font-weight: bold;
        color: #1a2a6c;
    }
    QLabel#SectionTitle {
        font-size: 15px;
        font-weight: bold;
        color: #2c3e50;
        border-bottom: 2px solid #3498db;
        padding-bottom: 4px;
    }
    QLineEdit {
        padding: 8px 10px;
        border: 2px solid #ddd;
        border-radius: 8px;
        background: #fff;
    }
    QLineEdit:focus { border: 2px solid #3498db; }
    QLineEdit:read-only {
        background: #f0f4f8;
        color: #334155;
        border: 1px solid #cbd5e1;
    }
    QPushButton {
        background-color: #3498db;
        color: white;
        padding: 8px 16px;
        border-radius: 8px;
        font-weight: bold;
        border: none;
    }
    QPushButton:hover { background-color: #2980b9; }
    QPushButton:disabled { background-color: #94a3b8; }
    QPushButton#SecondaryBtn { background-color: #95a5a6; }
    QPushButton#SecondaryBtn:hover { background-color: #7f8c8d; }
    QPushButton#ActionBtn {
        background-color: #27ae60;
        font-size: 14px;
        min-height: 40px;
    }
    QPushButton#ActionBtn:hover { background-color: #219150; }
    QPushButton#ActionBtn:disabled { background-color: #94a3b8; }
    QPushButton#EmailBtn {
        background-color: #e67e22;
        font-size: 14px;
        min-height: 40px;
        font-weight: bold;
    }
    QPushButton#EmailBtn:hover { background-color: #ca6f1e; }
    QPushButton#EmailBtn:disabled { background-color: #94a3b8; }
    QTextEdit#LogArea {
        background-color: #10151c;
        color: #d8e2f0;
        font-family: 'Consolas', 'Courier New';
        border-radius: 8px;
        padding: 10px;
        font-size: 12px;
        border: 1px solid #2b3a4f;
    }
    QTabWidget::pane {
        border: 1px solid #e1e4e8;
        border-radius: 8px;
        background: #f5f7fa;
    }
    QTabBar::tab {
        background: #ecf0f1;
        padding: 9px 20px;
        margin-right: 2px;
        border-top-left-radius: 8px;
        border-top-right-radius: 8px;
        font-weight: bold;
    }
    QTabBar::tab:selected { background: white; color: #1a2a6c; }
    QProgressBar {
        border: 1px solid #ddd;
        border-radius: 5px;
        min-height: 8px;
        max-height: 8px;
    }
    QProgressBar::chunk { background-color: #27ae60; border-radius: 5px; }
"""

STATE_FILE_NAME = "voucher_state.json"
AGENTS_CONFIG_FILE_NAME = "agents_config.json"
AGENT_NOTES_FILE_NAME = "agent_notes.json"
AGENT_RESPONSES_FOLDER = "תגובות סוכנים"
NOTE_COLUMNS = ["הערות", "הערות סוכן", "תאריך עידכון הערת סוכן"]
AGENT_REQUIRED_COLUMNS = ["משתמש בגלבוע", "סניף", "מייל", "מנהל סניף", "מנהל תחום", "מייל יפה", "מייל אילנית"]
_APP_DIR = os.path.dirname(os.path.abspath(__file__))

# File logger — writes to voucher_tool.log next to the script
_LOG_FILE = os.path.join(_APP_DIR, "voucher_tool.log")
logging.basicConfig(
    filename=_LOG_FILE,
    level=logging.DEBUG,
    encoding="utf-8",
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
_logger = logging.getLogger("voucher_tool")


class ProcessingError(Exception):
    pass


def safe_text(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def normalize_identifier(value) -> str:
    text = safe_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text


def safe_datetime(value, *, dayfirst: bool = False) -> Optional[datetime]:
    text = safe_text(value)
    if not text:
        return None
    result = pd.to_datetime(text, errors="coerce", dayfirst=dayfirst)
    if pd.isna(result):
        return None
    return result.to_pydatetime()


def parse_booster_date_from_filename(input_path: str) -> datetime:
    file_name = os.path.basename(input_path)
    name_without_ext = os.path.splitext(file_name)[0]
    match = re.search(r"(\d{1,2})[\s\-_/]+([A-Za-z]{3,9})[\s\-_/]+(\d{4})", name_without_ext)
    if not match:
        raise ProcessingError("לא נמצא תאריך בשם קובץ BOOSTER. פורמט נדרש: 12 Apr 2026")

    day = int(match.group(1))
    month_text = match.group(2).strip().lower()
    year = int(match.group(3))
    month_map = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }
    month = month_map.get(month_text)
    if month is None:
        raise ProcessingError(f"חודש לא תקין בשם קובץ BOOSTER: {match.group(2)}")

    try:
        return datetime(year, month, day)
    except ValueError as exc:
        raise ProcessingError(f"תאריך לא תקין בשם קובץ BOOSTER: {match.group(0)}") from exc


def build_alert_fields(days_to_departure: Optional[int], is_private: bool) -> tuple[str, str, str, str]:
    if not is_private or days_to_departure is None:
        return "", "", "", ""

    if 8 <= days_to_departure <= 14:
        category = "התראה שבועיים לפני היציאה"
        body = (
            "הלקוחות הבאים אמורים לצאת מהארץ בשבועיים הקרובים וטרם נגבה בגינם תשלום ההזמנה, "
            "במידה ותוך שבוע לא תתבצע גבייה היא תבוטל לאלתר בכדי למנוע הפסד לחברה"
        )
        reason = "חוסר גבייה כשבועיים לפני היציאה"
    elif 0 <= days_to_departure <= 7:
        category = "עלול להתבטל"
        body = (
            "הזמנת הלקוח טרם נגבתה, ואמורה לצאת לפועל תוך שבוע – לכן היא מיועדת לביטול ותבוטל עד סוף היום"
        )
        reason = "חוסר גבייה לפני היציאה"
    else:
        return "", "", "", ""

    return category, body, category, reason


def clean_numeric_series(series: pd.Series) -> pd.Series:
    cleaned = series.astype(str).str.replace(",", "", regex=False).str.strip()
    cleaned = cleaned.replace("", pd.NA)
    cleaned = cleaned.replace("nan", pd.NA)
    cleaned = cleaned.replace("None", pd.NA)
    return pd.to_numeric(cleaned, errors="coerce")


def load_agents_config(config_path: str) -> dict:
    if not os.path.exists(config_path):
        return {}
    with open(config_path, "r", encoding="utf-8") as file:
        return json.load(file)


def save_agents_config(config_path: str, data: dict) -> None:
    with open(config_path, "w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=2)


def find_latest_output_file(output_dir: str, source_name: str) -> Optional[str]:
    if not output_dir or not os.path.exists(output_dir):
        return None
    candidates = [
        path
        for path in glob.glob(os.path.join(output_dir, "*.xlsx"))
        if not os.path.basename(path).startswith("~$")
        and source_name.lower() in os.path.basename(path).lower()
    ]
    if not candidates:
        return None
    return max(candidates, key=os.path.getmtime)


def load_agent_table(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=0)
    df.columns = [str(c).strip() for c in df.columns]
    missing = [c for c in AGENT_REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ProcessingError("עמודות חסרות בטבלת סוכנים: " + ", ".join(missing))
    df["משתמש בגלבוע"] = df["משתמש בגלבוע"].astype(str).str.strip()
    return df


def _add_email_target(df: pd.DataFrame, client_col: str, *, empty_means_private: bool = False) -> pd.DataFrame:
    """Re-derives email_target from saved report columns (stripped during export)."""
    def _target(row):
        client = safe_text(row.get(client_col, "")) if client_col in row.index else ""
        is_private = (not client) if empty_means_private else (client == "Direct Sale")
        days_raw = row.get("מס' ימים ליציאה", None)
        try:
            days = int(days_raw) if pd.notna(days_raw) else None
        except (ValueError, TypeError):
            days = None
        _, _, email_target, _ = build_alert_fields(days, is_private)
        return email_target
    df = df.copy()
    df["email_target"] = df.apply(_target, axis=1)
    return df


def build_agent_rows_map(
    output_dir: str,
    agent_df: pd.DataFrame,
    logger: Callable[[str], None],
) -> tuple[Dict[str, pd.DataFrame], pd.DataFrame]:
    agent_lookup = agent_df[AGENT_REQUIRED_COLUMNS].copy()
    agent_lookup["_join_key"] = agent_lookup["משתמש בגלבוע"].str.strip().str.upper()
    matched_keys = set(agent_lookup["_join_key"])
    frames = []
    unmatched_frames = []

    booster_file = find_latest_output_file(output_dir, "BOOSTER")
    if booster_file:
        logger(f"טוען BOOSTER: {os.path.basename(booster_file)}")
        try:
            b_df = pd.read_excel(booster_file, sheet_name="דוח מלא")
            if "User" in b_df.columns:
                b_df = _add_email_target(b_df, "Agent/C. Client")
                b_df["_join_key"] = b_df["User"].astype(str).str.strip().str.upper()
                matched = b_df.merge(agent_lookup, on="_join_key", how="inner")
                matched = matched.drop(columns=["_join_key"])
                frames.append(matched)
                unmatched_b = b_df[~b_df["_join_key"].isin(matched_keys)].copy()
                unmatched_b = unmatched_b.rename(columns={"_join_key": "_agent_key"})
                if not unmatched_b.empty:
                    unmatched_frames.append(unmatched_b)
        except Exception as exc:
            logger(f"שגיאה בטעינת BOOSTER: {exc}")

    gilboa_file = find_latest_output_file(output_dir, "GILBOA")
    if gilboa_file:
        logger(f"טוען GILBOA: {os.path.basename(gilboa_file)}")
        try:
            g_df = pd.read_excel(gilboa_file, sheet_name="דוח מלא")
            if "Clerk" in g_df.columns:
                g_df = _add_email_target(g_df, "C.Client", empty_means_private=True)
                g_df["_join_key"] = g_df["Clerk"].astype(str).str.strip().str.upper()
                matched = g_df.merge(agent_lookup, on="_join_key", how="inner")
                matched = matched.drop(columns=["_join_key"])
                frames.append(matched)
                unmatched_g = g_df[~g_df["_join_key"].isin(matched_keys)].copy()
                unmatched_g = unmatched_g.rename(columns={"_join_key": "_agent_key"})
                if not unmatched_g.empty:
                    unmatched_frames.append(unmatched_g)
        except Exception as exc:
            logger(f"שגיאה בטעינת GILBOA: {exc}")

    if not frames and not unmatched_frames:
        raise ProcessingError("לא נמצאו נתונים לשיוך בקבצי BOOSTER/GILBOA בתיקיית הפלט")

    result: Dict[str, pd.DataFrame] = {}
    if frames:
        combined = pd.concat(frames, ignore_index=True)
        no_email_mask = combined["מייל"].isna() | (combined["מייל"].astype(str).str.strip() == "")
        no_email_df = combined[no_email_mask].copy()
        if not no_email_df.empty:
            agent_key_col = "משתמש בגלבוע" if "משתמש בגלבוע" in no_email_df.columns else "User"
            no_email_df["_agent_key"] = no_email_df[agent_key_col].astype(str).str.strip().str.upper()
            unmatched_frames.append(no_email_df)
        for email_addr, group in combined[~no_email_mask].groupby("מייל"):
            email_str = safe_text(email_addr)
            if email_str:
                result[email_str] = group.reset_index(drop=True)
    unmatched_combined = pd.concat(unmatched_frames, ignore_index=True) if unmatched_frames else pd.DataFrame()
    return result, unmatched_combined


def create_outlook_draft(
    to_email: str,
    cc_emails: list[str],
    subject: str,
    html_body: str,
    *,
    subfolder_name: str = "",
    save_as_path: str = "",
    attachments: list[str] | None = None,
    high_importance: bool = False,
) -> None:
    try:
        import win32com.client
    except ImportError as exc:
        raise ProcessingError("חבילת pywin32 אינה מותקנת. הרץ: pip install pywin32") from exc
    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.To = to_email
    if cc_emails:
        mail.CC = "; ".join(cc_emails)
    mail.Subject = subject
    mail.HTMLBody = html_body
    if high_importance:
        mail.Importance = 2  # olImportanceHigh
    if attachments:
        for att_path in attachments:
            mail.Attachments.Add(att_path)
    if save_as_path:
        mail.SaveAs(save_as_path, 3)  # 3 = olMSG
    else:
        mail.Save()
        if subfolder_name:
            namespace = outlook.GetNamespace("MAPI")
            drafts = namespace.GetDefaultFolder(16)
            try:
                target_folder = drafts.Folders[subfolder_name]
            except Exception:
                target_folder = drafts.Folders.Add(subfolder_name)
            mail.Move(target_folder)


def _save_df_excel(df: pd.DataFrame, path: str, *, split_by_source: bool = True, keep_cols: set[str] | None = None) -> None:
    """Write df to Excel. If multiple sources (מקור column) exist each gets its own sheet
    with only the columns that have at least one non-empty value in that group."""
    _keep = set(keep_cols) if keep_cols else set()

    def _drop_empty_cols(frame: pd.DataFrame) -> pd.DataFrame:
        def has_value(col):
            return col.apply(lambda v: pd.notna(v) and str(v).strip() not in ("", "nan", "None", "NaT")).any()
        mask = frame.apply(has_value) | frame.columns.isin(_keep)
        return frame.loc[:, mask]

    if split_by_source and "מקור" in df.columns and df["מקור"].nunique() > 1:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for source, group in df.groupby("מקור"):
                _drop_empty_cols(group.reset_index(drop=True)).to_excel(writer, sheet_name=str(source), index=False)
    else:
        _drop_empty_cols(df.reset_index(drop=True)).to_excel(path, index=False, engine="openpyxl")


def _sanitize_filename_component(value: str) -> str:
    text = safe_text(value)
    if not text:
        return "agent"
    return re.sub(r'[<>:"/\\|?*]', "-", text)


def _prepare_sorted_open_orders(rows_df: pd.DataFrame, routing_cols: set[str]) -> tuple[pd.DataFrame, list[str]]:
    working = rows_df.copy()
    required_display_cols = {"הערות סוכן", "תאריך עידכון הערת סוכן"}
    for col in required_display_cols:
        if col not in working.columns:
            working[col] = ""
    category_col = "email_target" if "email_target" in working.columns else None
    if category_col:
        priority = {
            "עלול להתבטל": 0,
            "התראה שבועיים לפני היציאה": 1,
        }
        working["_priority"] = working[category_col].map(priority).fillna(2).astype(int)
        working = working.sort_values(by=["_priority"], kind="stable").reset_index(drop=True)
        categories = working[category_col].apply(safe_text).tolist()
    else:
        working = working.reset_index(drop=True)
        categories = [""] * len(working)

    note_cols_order = ["הערות סוכן", "תאריך עידכון הערת סוכן"]
    display_cols = [
        c for c in working.columns
        if (c not in routing_cols and c != "_priority") and c not in required_display_cols
    ]
    display_cols = [c for c in display_cols if c in working.columns]
    for col in note_cols_order:
        if col in working.columns:
            display_cols.append(col)
    return working[display_cols].reset_index(drop=True), categories


def _apply_open_orders_excel_presentation(
    excel_path: str,
    categories: list[str],
    numeric_columns: list[str],
) -> None:
    date_columns = {
        "Open",
        "Start",
        "Start Date",
        "End Date",
        "Open Date",
        "Value Date",
        "תאריך עידכון הערת סוכן",
    }

    workbook = load_workbook(excel_path)
    worksheet = cast(Worksheet, workbook.active)
    headers = {cell.value: index + 1 for index, cell in enumerate(worksheet[1])}

    # Apply category row styles FIRST — named styles overwrite number_format,
    # so date/amount formats must be applied afterwards.
    style_by_category = {
        "עלול להתבטל": "Bad",
        "התראה שבועיים לפני היציאה": "Neutral",
    }
    for row_offset, category in enumerate(categories, start=2):
        row_style = style_by_category.get(safe_text(category))
        if not row_style:
            continue
        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row=row_offset, column=col).style = row_style

    for column_name in date_columns:
        col_idx = headers.get(column_name)
        if not col_idx:
            continue
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=col_idx).number_format = "YYYY-MM-DD"

    amount_columns = {"Balance", "Ref wl", "Credit", "Unconfirmed Refund"}
    for column_name in numeric_columns:
        if column_name in date_columns:
            continue
        col_idx = headers.get(column_name)
        if not col_idx:
            continue
        fmt = "#,##0.00" if column_name in amount_columns else "0"
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=col_idx).number_format = fmt

    workbook.save(excel_path)


def _build_open_orders_attachment(
    rows_df: pd.DataFrame,
    display_name: str,
    routing_cols: set[str],
) -> tuple[str, list[str], int, int, int]:
    sorted_df, categories = _prepare_sorted_open_orders(rows_df, routing_cols)
    category_series = rows_df["email_target"].apply(safe_text) if "email_target" in rows_df.columns else pd.Series(dtype=str)

    count_all = len(sorted_df)
    count_warning = int((category_series == "התראה שבועיים לפני היציאה").sum()) if not category_series.empty else 0
    count_cancel = int((category_series == "עלול להתבטל").sum()) if not category_series.empty else 0

    tmp_dir = tempfile.mkdtemp()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    safe_name = _sanitize_filename_component(display_name)
    all_path = os.path.join(tmp_dir, f"הזמנות פתוחות - {safe_name} {timestamp}.xlsx")

    _save_df_excel(sorted_df, all_path, split_by_source=False, keep_cols={"הערות סוכן", "תאריך עידכון הערת סוכן"})
    numeric_columns = sorted_df.select_dtypes(include="number").columns.tolist()
    _apply_open_orders_excel_presentation(all_path, categories, numeric_columns)

    return tmp_dir, [all_path], count_all, count_warning, count_cancel


def prepare_agent_emails(
    output_dir: str,
    agent_table_path: str,
    logger: Callable[[str], None],
) -> int:
    _ROUTING_COLS = set(AGENT_REQUIRED_COLUMNS) | {
        "email_target",
        "גוף דוא\u05f4ל",
    }
    _UNMATCHED_SUBFOLDER = "ממתינות מייל - כתובת חסרה"
    agent_df = load_agent_table(agent_table_path)
    rows_map, unmatched_df = build_agent_rows_map(output_dir, agent_df, logger)

    today_str = datetime.now().strftime("%d/%m/%Y")
    cc_source_columns = ["מנהל סניף", "מנהל תחום", "מייל יפה", "מייל אילנית"]
    count = 0
    for to_email, rows_df in rows_map.items():
        agent_name = safe_text(rows_df["משתמש בגלבוע"].iloc[0]) if "משתמש בגלבוע" in rows_df.columns else to_email
        cc_emails = []
        for col in cc_source_columns:
            if col in rows_df.columns:
                val = safe_text(rows_df[col].iloc[0])
                if val and val != to_email and val not in cc_emails:
                    cc_emails.append(val)

        tmp_dir, tmp_files, count_all, count_warning, count_cancel = _build_open_orders_attachment(
            rows_df,
            agent_name,
            _ROUTING_COLS,
        )
        _email_body_col = "גוף דוא״ל"
        body_lines = [
            "<p>שלום,</p>",
            "<p>להלן סיכום ההזמנות הפתוחות עבורך (הפרטים המלאים מצורפים כקבצי אקסל):</p>",
            f"<p>&#128196; <strong>הזמנות פתוחות:</strong> {count_all} רשומות</p>",
        ]
        if count_warning:
            body_lines.append(f"<p>&#9888; <strong>הזמנות שבועיים לפני היציאה:</strong> {count_warning} רשומות</p>")
        if count_cancel:
            body_lines.append(f"<p>&#128308; <strong>הזמנות עלולות להתבטל עד סוף היום:</strong> {count_cancel} רשומות</p>")
        if _email_body_col in rows_df.columns:
            body_counts = (
                rows_df[_email_body_col]
                .dropna()
                .loc[lambda s: s.str.strip() != ""]
                .value_counts()
            )
            for body_val, body_cnt in body_counts.items():
                body_lines.append(f"<p>{body_val} ({body_cnt})</p>")
        html_body = '<html><head></head><body dir="rtl">' + "".join(body_lines) + "</body></html>"

        subject = f"דוח בקרת דוקטים לגבייה \u2014 {agent_name} \u2014 {today_str}"
        create_outlook_draft(to_email, cc_emails, subject, html_body, attachments=tmp_files, high_importance=bool(count_warning or count_cancel))
        logger(f"נוצרה טיוטה עבור: {agent_name} ({to_email})")
        count += 1
        for p in tmp_files:
            try:
                os.remove(p)
            except Exception:
                pass
        try:
            os.rmdir(tmp_dir)
        except Exception:
            pass

    if not unmatched_df.empty:
        unmatched_out_dir = os.path.join(output_dir, _UNMATCHED_SUBFOLDER)
        os.makedirs(unmatched_out_dir, exist_ok=True)
        logger(f"נמצאו {len(unmatched_df)} רשומות ללא כתובת מייל — שומר קבצי מייל בתיקייה '{_UNMATCHED_SUBFOLDER}'")
        display_routing_exclude = _ROUTING_COLS | {"_agent_key"}
        for agent_key, grp in unmatched_df.groupby("_agent_key"):
            agent_key_str = safe_text(agent_key)
            tmp_dir, tmp_files, count_all, count_warning, count_cancel = _build_open_orders_attachment(
                grp,
                agent_key_str,
                display_routing_exclude,
            )
            _email_body_col = "גוף דוא״ל"
            body_lines = [
                f"<p style=\"color:#e11d48;\"><strong>&#9888; לא נמצאה כתובת מייל עבור סוכן: {agent_key_str}</strong></p>",
                "<p>יש להוסיף את כתובת המייל ידנית לפני שליחה.</p>",
                f"<p>&#128196; <strong>הזמנות פתוחות:</strong> {count_all} רשומות</p>",
            ]
            if count_warning:
                body_lines.append(f"<p>&#9888; <strong>הזמנות שבועיים לפני היציאה:</strong> {count_warning} רשומות</p>")
            if count_cancel:
                body_lines.append(f"<p>&#128308; <strong>הזמנות עלולות להתבטל עד סוף היום:</strong> {count_cancel} רשומות</p>")
            if _email_body_col in grp.columns:
                body_counts = (
                    grp[_email_body_col]
                    .dropna()
                    .loc[lambda s: s.str.strip() != ""]
                    .value_counts()
                )
                for body_val, body_cnt in body_counts.items():
                    body_lines.append(f"<p>{body_val} ({body_cnt})</p>")
            html_body = '<html><head></head><body dir="rtl">' + "".join(body_lines) + "</body></html>"

            subject = f"דוח בקרת דוקטים לגבייה \u2014 {agent_key_str} \u2014 {today_str} [חסרה כתובת מייל]"
            safe_key = agent_key_str.replace("/", "-").replace("\\", "-")
            msg_path = os.path.join(unmatched_out_dir, f"דוח {safe_key} {today_str.replace('/', '.')}.msg")
            _UNMATCHED_CC = ["ilanit_b@ophirtours.co.il", "YWaksman@mycwt.co.il"]
            create_outlook_draft("", _UNMATCHED_CC, subject, html_body, save_as_path=msg_path, attachments=tmp_files, high_importance=bool(count_warning or count_cancel))
            logger(f"קובץ מייל נשמר: {os.path.basename(msg_path)}")
            count += 1
            for p in tmp_files:
                try:
                    os.remove(p)
                except Exception:
                    pass
            try:
                os.rmdir(tmp_dir)
            except Exception:
                pass

    return count


def import_agent_responses(output_dir: str, logger: Callable[[str], None]) -> int:
    responses_dir = os.path.join(output_dir, AGENT_RESPONSES_FOLDER)
    if not os.path.exists(responses_dir):
        os.makedirs(responses_dir, exist_ok=True)
        logger(f"נוצרה תיקיית תגובות סוכנים: {responses_dir}")
        logger("התיקייה ריקה — יש להכניס קבצי Excel עם תגובות הסוכנים ולהפעיל שוב")
        return 0

    notes_path = os.path.join(output_dir, AGENT_NOTES_FILE_NAME)
    if os.path.exists(notes_path):
        with open(notes_path, "r", encoding="utf-8") as _f:
            agent_notes: dict = json.load(_f)
    else:
        agent_notes = {}

    xlsx_files = [
        p for p in glob.glob(os.path.join(responses_dir, "*.xlsx"))
        if not os.path.basename(p).startswith("~$")
    ]
    if not xlsx_files:
        raise ProcessingError(f"לא נמצאו קבצי Excel בתיקייה: {responses_dir}")

    imported = 0
    for file_path in xlsx_files:
        logger(f"קורא: {os.path.basename(file_path)}")
        try:
            df = pd.read_excel(file_path)
        except Exception as exc:
            logger(f"שגיאה בקריאת {os.path.basename(file_path)}: {exc}")
            continue
        df.columns = [str(c).strip() for c in df.columns]
        if "מזהה רשומה" not in df.columns or "מקור" not in df.columns:
            logger(f"עמודות 'מקור' / 'מזהה רשומה' חסרות בקובץ {os.path.basename(file_path)} — מדלג")
            continue
        for _, row in df.iterrows():
            source = safe_text(row.get("מקור", ""))
            record_id = normalize_identifier(row.get("מזהה רשומה", ""))
            note = safe_text(row.get("הערות סוכן", ""))
            date_str = safe_text(row.get("תאריך עידכון הערת סוכן", ""))
            if not source or not record_id or not note:
                continue
            if source not in agent_notes:
                agent_notes[source] = {}
            agent_notes[source][record_id] = {
                "הערות סוכן": note,
                "תאריך עידכון הערת סוכן": date_str,
            }
            imported += 1

    with open(notes_path, "w", encoding="utf-8") as _f:
        json.dump(agent_notes, _f, ensure_ascii=False, indent=2)
    logger(f"נקלטו {imported} הערות סוכנים ונשמרו ב-{AGENT_NOTES_FILE_NAME}")

    _patch_reports_with_notes(output_dir, agent_notes, logger)

    return imported


def _patch_reports_with_notes(
    output_dir: str,
    agent_notes: dict,
    logger: Callable[[str], None],
) -> None:
    """Update the latest Excel report for each source with agent notes in-place."""
    from openpyxl import load_workbook  # already a dependency

    NOTE_COL = "הערות סוכן"
    DATE_COL = "תאריך עידכון הערת סוכן"
    ID_COL = "מזהה רשומה"

    for source, notes_map in agent_notes.items():
        if not notes_map:
            continue
        report_path = find_latest_output_file(output_dir, source)
        if not report_path:
            logger(f"לא נמצא קובץ דוח עבור {source} — מדלג על עדכון הדוח")
            continue

        try:
            wb = load_workbook(report_path)
        except Exception as exc:
            logger(f"שגיאה בפתיחת {os.path.basename(report_path)}: {exc}")
            continue

        total_updated = 0
        for ws in wb.worksheets:
            # Read header row (row 1) to find column indices
            header = {
                str(ws.cell(1, col).value).strip(): col
                for col in range(1, ws.max_column + 1)
                if ws.cell(1, col).value is not None
            }
            id_col_idx = header.get(ID_COL)
            note_col_idx = header.get(NOTE_COL)
            date_col_idx = header.get(DATE_COL)

            if id_col_idx is None:
                continue  # sheet has no record-id column; skip

            sheet_updated = 0
            for row_idx in range(2, ws.max_row + 1):
                raw_id = ws.cell(row_idx, id_col_idx).value
                if raw_id is None:
                    continue
                record_id = normalize_identifier(raw_id)
                if record_id not in notes_map:
                    continue
                entry = notes_map[record_id]
                if note_col_idx is not None:
                    ws.cell(row_idx, note_col_idx).value = entry.get(NOTE_COL, "")
                if date_col_idx is not None:
                    ws.cell(row_idx, date_col_idx).value = entry.get(DATE_COL, "")
                sheet_updated += 1

            total_updated += sheet_updated

        try:
            wb.save(report_path)
            logger(f"עודכנו {total_updated} רשומות בדוח {os.path.basename(report_path)}")
        except Exception as exc:
            logger(f"שגיאה בשמירת {os.path.basename(report_path)}: {exc}")


class BaseVoucherProcessor:
    source_name = "BASE"
    history_key_columns = []

    def __init__(self, logger: Optional[Callable[[str], None]] = None):
        self.logger = logger or (lambda message: None)

    def log(self, message: str) -> None:
        self.logger(message)

    def ensure_columns(self, df: pd.DataFrame, required_columns: list[str]) -> None:
        missing = [column for column in required_columns if column not in df.columns]
        if missing:
            raise ProcessingError("עמודות חסרות בקובץ: " + ", ".join(missing))

    def load_state(self, state_path: str) -> Dict[str, Dict[str, str]]:
        if not os.path.exists(state_path):
            return {}
        with open(state_path, "r", encoding="utf-8") as file:
            return json.load(file)

    def save_state(self, state_path: str, state: Dict[str, Dict[str, str]]) -> None:
        with open(state_path, "w", encoding="utf-8") as file:
            json.dump(state, file, ensure_ascii=False, indent=2)

    def get_anchor_date(self, input_path: str) -> datetime:
        return datetime.now()

    def build_history_map(self, output_dir: str) -> Dict[str, Dict[str, str]]:
        if not output_dir or not os.path.exists(output_dir):
            return {}

        candidates = []
        for pattern in ("*.xlsx", "*.csv"):
            candidates.extend(
                [
                    path
                    for path in glob.glob(os.path.join(output_dir, pattern))
                    if not os.path.basename(path).startswith("~$")
                ]
            )

        if not candidates:
            return {}

        source_matches = [
            path for path in candidates if self.source_name.lower() in os.path.basename(path).lower()
        ]
        selected = source_matches if source_matches else candidates
        latest_file = max(selected, key=os.path.getmtime)
        self.log(f"טוען היסטוריית הערות: {os.path.basename(latest_file)}")

        try:
            if latest_file.lower().endswith(".xlsx"):
                old_df = pd.read_excel(latest_file)
            else:
                old_df = pd.read_csv(latest_file, encoding="utf-8-sig")
        except Exception:
            return {}

        id_column = None
        if "מקור" in old_df.columns and "מזהה רשומה" in old_df.columns:
            old_df = old_df[old_df["מקור"].astype(str).str.upper() == self.source_name.upper()].copy()
            id_column = "מזהה רשומה"
        else:
            for candidate in self.history_key_columns:
                if candidate in old_df.columns:
                    id_column = candidate
                    break

        if not id_column:
            return {}

        map_data: Dict[str, Dict[str, str]] = {}
        for _, row in old_df.iterrows():
            record_id = normalize_identifier(row.get(id_column, ""))
            if not record_id:
                continue
            map_data[record_id] = {
                "הערות סוכן": safe_text(row.get("הערות סוכן", "")),
                "תאריך עידכון הערת סוכן": safe_text(row.get("תאריך עידכון הערת סוכן", "")),
            }

        # Overlay notes that agents filled and returned via the responses folder
        notes_path = os.path.join(output_dir, AGENT_NOTES_FILE_NAME)
        if os.path.exists(notes_path):
            try:
                with open(notes_path, "r", encoding="utf-8") as _f:
                    agent_notes = json.load(_f)
                for rec_id, notes in agent_notes.get(self.source_name, {}).items():
                    note_text = notes.get("הערות סוכן", "")
                    if note_text:
                        if rec_id not in map_data:
                            map_data[rec_id] = {}
                        map_data[rec_id]["הערות סוכן"] = note_text
                        map_data[rec_id]["תאריך עידכון הערת סוכן"] = notes.get("תאריך עידכון הערת סוכן", "")
            except Exception:
                pass

        return map_data

    def history_lookup(self, history_map: Dict[str, Dict[str, str]], raw_identifier) -> Dict[str, str]:
        record_id = normalize_identifier(raw_identifier)
        return history_map.get(record_id, {})

    def export_report(self, df: pd.DataFrame, output_dir: str) -> str:
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%d.%m.%y %H.%M")
        output_path = os.path.join(output_dir, f"דוח בקרת דוקטים לגבייה {self.source_name} {timestamp}.xlsx")

        warning_14 = pd.DataFrame()
        threat_7 = pd.DataFrame()
        if "קטגוריית התראה" in df.columns:
            warning_14 = df[df["קטגוריית התראה"] == "התראה שבועיים לפני היציאה"].copy()
            threat_7 = df[df["קטגוריית התראה"] == "עלול להתבטל"].copy()

        drop_columns = ["קטגוריית התראה", "סיבת התראה", "גוף דוא״ל", "email_target"]
        if any(column in df.columns for column in drop_columns):
            df = df.drop(columns=drop_columns, errors="ignore")
            warning_14 = warning_14.drop(columns=drop_columns, errors="ignore")
            threat_7 = threat_7.drop(columns=drop_columns, errors="ignore")

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="דוח מלא", index=False)
            if not warning_14.empty:
                warning_14.to_excel(writer, sheet_name="אזהרה שבועיים", index=False)
            if not threat_7.empty:
                threat_7.to_excel(writer, sheet_name="עלול להתבטל", index=False)

        workbook = load_workbook(output_path)
        date_columns = {
            "Open",
            "Start",
            "Start Date",
            "End Date",
            "Open Date",
            "Value Date",
            "תאריך עידכון הערת סוכן",
        }
        integer_columns = {"Number", "מס' ימים ליציאה", "מס' ימי החזר"}
        decimal_columns = {"Ref wl", "Attr", "Credit", "Balance", "Unconfirmed Refund"}

        for worksheet in workbook.worksheets:
            headers = {cell.value: index + 1 for index, cell in enumerate(worksheet[1])}
            for column_name in date_columns:
                col_idx = headers.get(column_name)
                if not col_idx:
                    continue
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=col_idx).number_format = "DD/MM/YYYY"
            for column_name in integer_columns:
                col_idx = headers.get(column_name)
                if not col_idx:
                    continue
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=col_idx).number_format = "0"
            for column_name in decimal_columns:
                col_idx = headers.get(column_name)
                if not col_idx:
                    continue
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=col_idx).number_format = "#,##0.00"

        workbook.save(output_path)
        return output_path

    def process(self, input_path: str, output_dir: str) -> str:
        self.log(f"מתחיל עיבוד עבור {self.source_name}")
        self.log(f"קובץ קלט: {os.path.basename(input_path)}")

        state_path = os.path.join(output_dir, STATE_FILE_NAME)
        state = self.load_state(state_path)
        history = self.build_history_map(output_dir)
        anchor_date = self.get_anchor_date(input_path)

        df = self.load_input(input_path)
        self.log(f"נטענו {len(df)} רשומות")
        result_df = self.apply_business_logic(df, state, history, anchor_date)
        self.save_state(state_path, state)
        output_path = self.export_report(result_df, output_dir)
        self.log(f"הפלט נשמר אל: {output_path}")
        return output_path

    def load_input(self, input_path: str) -> pd.DataFrame:
        raise NotImplementedError

    def apply_business_logic(
        self,
        df: pd.DataFrame,
        state: Dict[str, Dict[str, str]],
        history: Dict[str, Dict[str, str]],
        anchor_date: datetime,
    ) -> pd.DataFrame:
        raise NotImplementedError


class BoosterProcessor(BaseVoucherProcessor):
    source_name = "BOOSTER"
    history_key_columns = ["מזהה רשומה", "T. File No."]

    def get_anchor_date(self, input_path: str) -> datetime:
        anchor = parse_booster_date_from_filename(input_path)
        self.log(f"תאריך העוגן של BOOSTER נקבע משם הקובץ: {anchor.strftime('%d/%m/%Y')}")
        return anchor

    def load_input(self, input_path: str) -> pd.DataFrame:
        self.log("קורא קובץ BOOSTER")
        df = pd.read_excel(input_path, sheet_name="Ext. T. File Report", skiprows=8)
        df.columns = [str(column).strip() for column in df.columns]
        self.ensure_columns(
            df,
            [
                "Branch",
                "T. File No.",
                "User",
                "Open Date",
                "Value Date",
                "Start Date",
                "End Date",
                "T. File Name",
                "Agent/C. Client",
                "Unconfirmed Refund",
                "Balance",
            ],
        )
        df = df[df["Branch"].notna()].copy()
        return df[
            [
                "T. File No.",
                "User",
                "Open Date",
                "Value Date",
                "Start Date",
                "End Date",
                "T. File Name",
                "Agent/C. Client",
                "Unconfirmed Refund",
                "Balance",
            ]
        ].copy()

    def apply_business_logic(
        self,
        df: pd.DataFrame,
        state: Dict[str, Dict[str, str]],
        history: Dict[str, Dict[str, str]],
        anchor_date: datetime,
    ) -> pd.DataFrame:
        df["Open Date"] = pd.to_datetime(df["Open Date"], errors="coerce")
        df["Value Date"] = pd.to_datetime(df["Value Date"], errors="coerce")
        df["Start Date"] = pd.to_datetime(df["Start Date"], errors="coerce")
        df["User"] = df["User"].astype(str).str.strip()
        df["מקור"] = self.source_name
        df["מזהה רשומה"] = df["T. File No."].apply(normalize_identifier)

        run_date = anchor_date.date()
        calc_columns = []
        current_record_ids = set()
        positive_refund_ids = set()
        for _, row in df.iterrows():
            record_id = normalize_identifier(row["T. File No."])
            if record_id:
                current_record_ids.add(record_id)
            start_date = row["Start Date"] if pd.notna(row["Start Date"]) else None
            state_key = f"{self.source_name}:{record_id}" if record_id else ""

            start_date_only = start_date.date() if start_date is not None else None
            days_to_departure = (
                int((start_date_only - run_date).days) if start_date_only is not None else None
            )

            refund_value = pd.to_numeric(row.get("Unconfirmed Refund", 0), errors="coerce")
            refund_value = 0 if pd.isna(refund_value) else float(refund_value)
            refund_positive = refund_value > 0

            if refund_positive and record_id:
                positive_refund_ids.add(record_id)
                state_entry = state.get(state_key, {})
                recognition_date = safe_datetime(state_entry.get("open_date", ""))
                if recognition_date is None:
                    state[state_key] = {"open_date": str(anchor_date)}
                    recognition_date = anchor_date
                days_since_refund = (anchor_date - recognition_date).days
            else:
                days_since_refund = 0

            refund_note = ""
            if refund_positive:
                refund_note = "החזר שטרם אושר מעל 60 יום" if days_since_refund > 60 else "החזר שטרם אושר"

            is_private = safe_text(row["Agent/C. Client"]) == "Direct Sale"
            alert_category, email_body, email_target, alert_reason = build_alert_fields(
                days_to_departure, is_private
            )

            if safe_text(row["Agent/C. Client"]) == "Direct Sale":
                system_note = refund_note
            else:
                system_note = f"{refund_note}; טרם הופקה חשבונית" if refund_note else "טרם הופקה חשבונית"

            hist = self.history_lookup(history, row["T. File No."])
            calc_columns.append(
                [
                    days_to_departure,
                    days_since_refund,
                    system_note,
                    hist.get("הערות סוכן", ""),
                    hist.get("תאריך עידכון הערת סוכן", ""),
                    alert_category,
                    email_body,
                    email_target,
                    alert_reason,
                ]
            )

        booster_prefix = f"{self.source_name}:"
        removed_missing = 0
        removed_non_positive = 0
        for state_key in [k for k in state.keys() if k.startswith(booster_prefix)]:
            record_id = state_key[len(booster_prefix):]
            if record_id not in current_record_ids:
                del state[state_key]
                removed_missing += 1
            elif record_id not in positive_refund_ids:
                del state[state_key]
                removed_non_positive += 1
        if removed_missing or removed_non_positive:
            self.log(
                "ניקוי JSON BOOSTER: "
                f"נמחקו {removed_missing} רשומות שלא הופיעו בקובץ ו-{removed_non_positive} רשומות ללא החזר חיובי"
            )

        df[
            [
                "מס' ימים ליציאה",
                "מס' ימי החזר",
                "הערות",
                "הערות סוכן",
                "תאריך עידכון הערת סוכן",
                "קטגוריית התראה",
                "גוף דוא״ל",
                "email_target",
                "סיבת התראה",
            ]
        ] = pd.DataFrame(calc_columns, index=df.index)

        preferred = [
            "מקור",
            "מזהה רשומה",
            "T. File No.",
            "User",
            "Open Date",
            "Value Date",
            "Start Date",
            "End Date",
            "T. File Name",
            "Agent/C. Client",
            "Unconfirmed Refund",
            "Balance",
            "מס' ימים ליציאה",
            "מס' ימי החזר",
            "הערות",
            "הערות סוכן",
            "תאריך עידכון הערת סוכן",
            "קטגוריית התראה",
            "סיבת התראה",
            "גוף דוא״ל",
            "email_target",
        ]
        return df[preferred]


class GilboaProcessor(BaseVoucherProcessor):
    source_name = "GILBOA"
    history_key_columns = ["מזהה רשומה", "Number"]

    def get_anchor_date(self, input_path: str) -> datetime:
        anchor = datetime.fromtimestamp(os.path.getmtime(input_path))
        self.log(f"תאריך העוגן של GILBOA נקבע ממועד שינוי הקובץ: {anchor.strftime('%d/%m/%Y')}")
        return anchor

    def load_input(self, input_path: str) -> pd.DataFrame:
        self.log("קורא קובץ GILBOA")
        read_errors = []
        for encoding in ("cp1255", "utf-8-sig", "utf-8"):
            try:
                df = pd.read_csv(
                    input_path,
                    sep=r"\s*\^\s*",
                    engine="python",
                    skipinitialspace=True,
                    encoding=encoding,
                )
                df = df.loc[:, ~df.columns.str.contains("^Unnamed")]
                df.columns = [str(column).strip() for column in df.columns]
                self.ensure_columns(df, ["Number", "Open", "Start", "Ref wl", "C.Client", "Clerk"])
                return df
            except Exception as error:
                read_errors.append(str(error))

        raise ProcessingError("לא ניתן לקרוא את קובץ GILBOA. " + " | ".join(read_errors))

    def apply_business_logic(
        self,
        df: pd.DataFrame,
        state: Dict[str, Dict[str, str]],
        history: Dict[str, Dict[str, str]],
        anchor_date: datetime,
    ) -> pd.DataFrame:
        df["Open"] = pd.to_datetime(df["Open"], format="%d/%m/%Y", errors="coerce")
        df["Start"] = pd.to_datetime(df["Start"], format="%d/%m/%Y", errors="coerce")
        df["Clerk"] = df["Clerk"].astype(str).str.strip()

        for numeric_column in ["Number", "Ref wl", "Attr", "Credit", "Balance"]:
            if numeric_column in df.columns:
                df[numeric_column] = clean_numeric_series(df[numeric_column])

        df["מקור"] = self.source_name
        df["מזהה רשומה"] = df["Number"].apply(normalize_identifier)

        run_date = anchor_date.date()
        calc_columns = []
        current_record_ids = set()
        nonzero_refund_ids = set()
        for _, row in df.iterrows():
            record_id = normalize_identifier(row["Number"])
            if record_id:
                current_record_ids.add(record_id)
            start_date = row["Start"].to_pydatetime() if pd.notna(row["Start"]) else None
            state_key = f"{self.source_name}:{record_id}" if record_id else ""

            start_date_only = start_date.date() if start_date is not None else None
            days_to_departure = (
                int((start_date_only - run_date).days) if start_date_only is not None else None
            )

            refund_value = row.get("Ref wl", 0)
            refund_value = 0.0 if pd.isna(refund_value) else float(refund_value)

            if refund_value != 0 and record_id:
                nonzero_refund_ids.add(record_id)
                state_entry = state.get(state_key, {})
                recognition_date = safe_datetime(state_entry.get("open_date", ""))
                if recognition_date is None:
                    state[state_key] = {"open_date": str(anchor_date)}
                    recognition_date = anchor_date
                days_since_refund = (anchor_date - recognition_date).days
            else:
                days_since_refund = 0

            refund_note = ""
            if refund_value != 0:
                refund_note = "החזר שטרם אושר מעל 60 יום" if days_since_refund > 60 else "החזר שטרם אושר"

            is_private = not safe_text(row.get("C.Client", ""))
            alert_category, email_body, email_target, alert_reason = build_alert_fields(
                days_to_departure, is_private
            )

            client_text = safe_text(row.get("C.Client", ""))
            has_invoice_gap = bool(client_text)
            if has_invoice_gap and refund_note:
                system_note = f"{refund_note}; טרם הופקה חשבונית"
            elif has_invoice_gap:
                system_note = "טרם הופקה חשבונית"
            else:
                system_note = refund_note

            hist = self.history_lookup(history, row["Number"])
            calc_columns.append(
                [
                    days_to_departure,
                    days_since_refund,
                    system_note,
                    hist.get("הערות סוכן", ""),
                    hist.get("תאריך עידכון הערת סוכן", ""),
                    alert_category,
                    email_body,
                    email_target,
                    alert_reason,
                ]
            )

        gilboa_prefix = f"{self.source_name}:"
        removed_missing = 0
        removed_non_refund = 0
        for state_key in [k for k in state.keys() if k.startswith(gilboa_prefix)]:
            record_id = state_key[len(gilboa_prefix):]
            if record_id not in current_record_ids:
                del state[state_key]
                removed_missing += 1
            elif record_id not in nonzero_refund_ids:
                del state[state_key]
                removed_non_refund += 1
        if removed_missing or removed_non_refund:
            self.log(
                "ניקוי JSON GILBOA: "
                f"נמחקו {removed_missing} רשומות שלא הופיעו בקובץ ו-{removed_non_refund} רשומות ללא החזר"
            )

        df[
            [
                "מס' ימים ליציאה",
                "מס' ימי החזר",
                "הערות",
                "הערות סוכן",
                "תאריך עידכון הערת סוכן",
                "קטגוריית התראה",
                "גוף דוא״ל",
                "email_target",
                "סיבת התראה",
            ]
        ] = pd.DataFrame(calc_columns, index=df.index)

        added_columns = [
            "מס' ימים ליציאה",
            "מס' ימי החזר",
            "הערות",
            "הערות סוכן",
            "תאריך עידכון הערת סוכן",
            "קטגוריית התראה",
            "סיבת התראה",
            "גוף דוא״ל",
            "email_target",
        ]
        leading_columns = [
            "מקור",
            "מזהה רשומה",
            "Number",
            "Open",
            "Start",
            "Ref wl",
            "C.Client",
            "Clerk",
        ]
        middle_columns = [
            column for column in df.columns if column not in leading_columns and column not in added_columns
        ]
        return df[leading_columns + middle_columns + added_columns]


class StatCard(QFrame):
    """Small summary card showing a count and a label."""

    _STYLE_MATCHED = (
        "QFrame { background: #dcfce7; border: 1.5px solid #16a34a; border-radius: 10px; }"
    )
    _STYLE_UNMATCHED = (
        "QFrame { background: #fff7ed; border: 1.5px solid #ea580c; border-radius: 10px; }"
    )
    _STYLE_IDLE = (
        "QFrame { background: #f1f5f9; border: 1.5px solid #cbd5e1; border-radius: 10px; }"
    )

    def __init__(self, variant: str, parent=None):
        super().__init__(parent)
        self._variant = variant  # "matched" | "unmatched"
        self.setMinimumHeight(72)

        inner = QVBoxLayout(self)
        inner.setContentsMargins(14, 8, 14, 8)
        inner.setSpacing(2)

        self._number_label = QLabel("—")
        num_font = QFont("Arial", 26, QFont.Weight.Bold)
        self._number_label.setFont(num_font)
        self._number_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self._title_label = QLabel()
        title_font = QFont("Arial", 9)
        self._title_label.setFont(title_font)
        self._title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self._sub_label = QLabel()
        sub_font = QFont("Arial", 8)
        self._sub_label.setFont(sub_font)
        self._sub_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        inner.addWidget(self._number_label)
        inner.addWidget(self._title_label)
        inner.addWidget(self._sub_label)

        self._set_idle()

    def _set_idle(self) -> None:
        self.setStyleSheet(self._STYLE_IDLE)
        self._number_label.setStyleSheet("color: #94a3b8;")
        self._title_label.setStyleSheet("color: #94a3b8;")
        self._sub_label.setStyleSheet("color: #94a3b8;")
        self._number_label.setText("—")
        self._title_label.setText(
            "סוכנים ששויכו" if self._variant == "matched" else "ללא כתובת מייל"
        )
        self._sub_label.setText("")

    def set_value(self, count: int, sub: str = "") -> None:
        if self._variant == "matched":
            self.setStyleSheet(self._STYLE_MATCHED)
            self._number_label.setStyleSheet("color: #15803d;")
            self._title_label.setStyleSheet("color: #166534;")
            self._sub_label.setStyleSheet("color: #166534;")
            self._title_label.setText("סוכנים ששויכו")
        else:
            self.setStyleSheet(self._STYLE_UNMATCHED if count > 0 else self._STYLE_IDLE)
            color = "#c2410c" if count > 0 else "#94a3b8"
            self._number_label.setStyleSheet(f"color: {color};")
            self._title_label.setStyleSheet(f"color: {color};")
            self._sub_label.setStyleSheet(f"color: {color};")
            self._title_label.setText("ללא כתובת מייל")
        self._number_label.setText(str(count))
        self._sub_label.setText(sub)

    def reset(self) -> None:
        self._set_idle()


class ProcessWorker(QObject):
    finished = Signal(str)
    error = Signal(str)
    log_message = Signal(str)

    def __init__(self, processor_cls, input_path: str, output_dir: str):
        super().__init__()
        self.processor_cls = processor_cls
        self.input_path = input_path
        self.output_dir = output_dir

    @Slot()
    def run(self) -> None:
        try:
            processor = self.processor_cls(logger=self.log_message.emit)
            output_path = processor.process(self.input_path, self.output_dir)
            self.finished.emit(output_path)
        except Exception as error:
            self.error.emit(str(error))


class EmailWorker(QObject):
    finished = Signal(int)
    error = Signal(str)
    log_message = Signal(str)

    def __init__(self, output_dir: str, agent_table_path: str):
        super().__init__()
        self.output_dir = output_dir
        self.agent_table_path = agent_table_path

    @Slot()
    def run(self) -> None:
        import pythoncom
        pythoncom.CoInitialize()
        try:
            count = prepare_agent_emails(self.output_dir, self.agent_table_path, self.log_message.emit)
            self.finished.emit(count)
        except Exception as error:
            self.error.emit(str(error))
        finally:
            pythoncom.CoUninitialize()


class ImportWorker(QObject):
    finished = Signal(int)
    error = Signal(str)
    log_message = Signal(str)

    def __init__(self, output_dir: str):
        super().__init__()
        self.output_dir = output_dir

    @Slot()
    def run(self) -> None:
        try:
            count = import_agent_responses(self.output_dir, self.log_message.emit)
            self.finished.emit(count)
        except Exception as error:
            self.error.emit(str(error))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.worker_thread: Optional[QThread] = None
        self.worker: Optional[ProcessWorker] = None
        self.email_worker_thread: Optional[QThread] = None
        self.email_worker: Optional[EmailWorker] = None
        self.import_worker_thread: Optional[QThread] = None
        self.import_worker: Optional[ImportWorker] = None
        self.output_dir = ""
        self.agent_table_path = ""

        _config = load_agents_config(os.path.join(_APP_DIR, AGENTS_CONFIG_FILE_NAME))
        _saved = _config.get("agent_table_path", "")
        if _saved and os.path.exists(_saved):
            self.agent_table_path = _saved

        self.setWindowTitle(APP_TITLE)
        self.resize(1000, 750)
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.setStyleSheet(MODERN_STYLE)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(16)

        # --- Header ---
        header_layout = QHBoxLayout()
        title_label = QLabel(APP_TITLE)
        title_label.setObjectName("HeaderLabel")
        header_layout.addWidget(title_label)
        header_layout.addStretch()
        date_label = QLabel(datetime.now().strftime("%d/%m/%Y"))
        date_label.setStyleSheet("color: #7f8c8d; font-weight: bold;")
        header_layout.addWidget(date_label)
        main_layout.addLayout(header_layout)

        # --- Tabs ---
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)

        # ===== MAIN TAB =====
        main_tab = QWidget()
        tab_layout = QVBoxLayout(main_tab)
        tab_layout.setSpacing(16)
        tab_layout.setContentsMargins(16, 16, 16, 16)

        # Output Folder Card
        output_card = QFrame()
        output_card.setObjectName("Card")
        output_card_layout = QVBoxLayout(output_card)
        output_card_layout.setContentsMargins(16, 12, 16, 12)
        sec1 = QLabel("📁 תיקיית פלט")
        sec1.setObjectName("SectionTitle")
        output_card_layout.addWidget(sec1)
        path_layout = QHBoxLayout()
        self.output_path_edit = QLineEdit()
        self.output_path_edit.setPlaceholderText("בחר תיקיית פלט לשמירת הדוחות וההיסטוריה")
        self.output_path_edit.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        browse_button = QPushButton("בחר תיקיית פלט")
        browse_button.setObjectName("SecondaryBtn")
        browse_button.clicked.connect(self.choose_output_dir)
        path_layout.addWidget(browse_button)
        path_layout.addWidget(self.output_path_edit, stretch=1)
        output_card_layout.addLayout(path_layout)
        tab_layout.addWidget(output_card)

        # Agent Table Card
        agent_card = QFrame()
        agent_card.setObjectName("Card")
        agent_card_layout = QVBoxLayout(agent_card)
        agent_card_layout.setContentsMargins(16, 12, 16, 12)
        sec2 = QLabel("👤 טבלת סוכנים")
        sec2.setObjectName("SectionTitle")
        agent_card_layout.addWidget(sec2)
        agent_layout = QHBoxLayout()
        self.agent_path_edit = QLineEdit()
        self.agent_path_edit.setReadOnly(True)
        self.agent_path_edit.setPlaceholderText("טבלת סוכנים לא נטענה")
        self.agent_path_edit.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        if self.agent_table_path:
            self.agent_path_edit.setText(self.agent_table_path)
        agent_browse_button = QPushButton("קליטת טבלת סוכנים")
        agent_browse_button.setObjectName("SecondaryBtn")
        agent_browse_button.clicked.connect(self.choose_agent_table)
        agent_layout.addWidget(agent_browse_button)
        agent_layout.addWidget(self.agent_path_edit, stretch=1)
        agent_card_layout.addLayout(agent_layout)
        tab_layout.addWidget(agent_card)

        # Actions Card
        actions_card = QFrame()
        actions_card.setObjectName("Card")
        actions_card_layout = QVBoxLayout(actions_card)
        actions_card_layout.setContentsMargins(16, 12, 16, 12)
        sec3 = QLabel("⚡ פעולות מהירות")
        sec3.setObjectName("SectionTitle")
        actions_card_layout.addWidget(sec3)
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(10)
        self.booster_button = QPushButton("קליטת קובץ בוסטר")
        self.booster_button.setObjectName("ActionBtn")
        self.gilboa_button = QPushButton("קליטת קובץ גילבוע")
        self.gilboa_button.setObjectName("ActionBtn")
        self.email_button = QPushButton("📧 הכנת מיילים לסוכנים")
        self.email_button.setObjectName("EmailBtn")
        self.import_button = QPushButton("קליטת תגובות סוכנים")
        self.import_button.setObjectName("ActionBtn")
        self.booster_button.clicked.connect(lambda: self.select_and_run(BoosterProcessor, "קבצי Excel (*.xlsx *.xls)"))
        self.gilboa_button.clicked.connect(lambda: self.select_and_run(GilboaProcessor, "קבצי טקסט (*.txt)"))
        self.email_button.clicked.connect(self.start_email_worker)
        self.import_button.clicked.connect(self.start_import_worker)
        buttons_layout.addWidget(self.booster_button)
        buttons_layout.addWidget(self.gilboa_button)
        buttons_layout.addWidget(self.email_button)
        buttons_layout.addWidget(self.import_button)
        actions_card_layout.addLayout(buttons_layout)
        tab_layout.addWidget(actions_card)

        # Stats Card
        stats_card = QFrame()
        stats_card.setObjectName("Card")
        stats_card_layout = QVBoxLayout(stats_card)
        stats_card_layout.setContentsMargins(16, 12, 16, 12)
        sec4 = QLabel("📊 סיכום שיוך סוכנים לכתובות מייל")
        sec4.setObjectName("SectionTitle")
        stats_card_layout.addWidget(sec4)
        stats_layout = QHBoxLayout()
        stats_layout.setSpacing(12)
        self.stat_unmatched = StatCard("unmatched")
        self.stat_matched = StatCard("matched")
        stats_layout.addWidget(self.stat_unmatched)
        stats_layout.addWidget(self.stat_matched)
        stats_card_layout.addLayout(stats_layout)
        tab_layout.addWidget(stats_card)

        # Progress Bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(False)
        tab_layout.addWidget(self.progress_bar)

        tab_layout.addStretch()
        self.tabs.addTab(main_tab, "מסך ראשי")

        # ===== LOGS TAB =====
        logs_tab = QWidget()
        logs_layout = QVBoxLayout(logs_tab)
        logs_layout.setContentsMargins(16, 16, 16, 16)
        logs_layout.setSpacing(8)
        self.log_view = QTextEdit()
        self.log_view.setReadOnly(True)
        self.log_view.setObjectName("LogArea")
        self.log_view.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        logs_layout.addWidget(self.log_view)
        clear_btn = QPushButton("נקה לוג")
        clear_btn.setObjectName("SecondaryBtn")
        clear_btn.setFixedWidth(100)
        clear_btn.clicked.connect(self.log_view.clear)
        logs_layout.addWidget(clear_btn, alignment=Qt.AlignmentFlag.AlignLeft)
        self.tabs.addTab(logs_tab, "לוג פעילות")

        # Status Bar
        self.statusBar().showMessage("מוכן לעבודה")

        if self.agent_table_path:
            self.append_log(f"טבלת סוכנים נטענה אוטומטית: {self.agent_table_path}")
        self.append_log("המערכת מוכנה לעבודה")

    def append_log(self, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_view.append(f"[{timestamp}] {message}")
        self.statusBar().showMessage(message)
        _logger.info(message)

    def choose_output_dir(self) -> None:
        selected_dir = QFileDialog.getExistingDirectory(self, "בחר תיקיית פלט")
        if selected_dir:
            self.output_dir = selected_dir
            self.output_path_edit.setText(selected_dir)
            self.append_log(f"נבחרה תיקיית פלט: {selected_dir}")
            responses_dir = os.path.join(selected_dir, AGENT_RESPONSES_FOLDER)
            if not os.path.exists(responses_dir):
                os.makedirs(responses_dir, exist_ok=True)
                self.append_log(f"נוצרה תיקיית תגובות סוכנים: {responses_dir}")

    def select_and_run(self, processor_cls, file_filter: str) -> None:
        output_dir = self.output_path_edit.text().strip()
        if not output_dir:
            QMessageBox.warning(self, APP_TITLE, "יש לבחור תיקיית פלט לפני תחילת העבודה")
            return

        file_path, _ = QFileDialog.getOpenFileName(self, "בחר קובץ לעיבוד", "", file_filter)
        if not file_path:
            return

        self.append_log(f"נבחר קובץ: {file_path}")
        self.start_worker(processor_cls, file_path, output_dir)

    def set_busy(self, is_busy: bool) -> None:
        self.booster_button.setDisabled(is_busy)
        self.gilboa_button.setDisabled(is_busy)
        self.email_button.setDisabled(is_busy)
        self.progress_bar.setVisible(is_busy)
        if is_busy:
            self.progress_bar.setRange(0, 0)
            self.statusBar().showMessage("מעבד...")
        else:
            self.progress_bar.setRange(0, 100)
            self.progress_bar.setValue(100)
            self.statusBar().showMessage("מוכן לעבודה")

    def start_worker(self, processor_cls, input_path: str, output_dir: str) -> None:
        if self.worker_thread is not None and self.worker_thread.isRunning():
            QMessageBox.information(self, APP_TITLE, "כבר מתבצע עיבוד, יש להמתין לסיום")
            return

        self.set_busy(True)
        self.worker_thread = QThread(self)
        self.worker = ProcessWorker(processor_cls, input_path, output_dir)
        self.worker.moveToThread(self.worker_thread)

        self.worker_thread.started.connect(self.worker.run)
        self.worker.log_message.connect(self.append_log)
        self.worker.finished.connect(self.on_finished)
        self.worker.error.connect(self.on_error)
        self.worker.finished.connect(self.worker_thread.quit)
        self.worker.error.connect(self.worker_thread.quit)
        self.worker_thread.finished.connect(self.cleanup_thread)
        self.worker_thread.start()

    def cleanup_thread(self) -> None:
        self.set_busy(False)
        if self.worker is not None:
            self.worker.deleteLater()
            self.worker = None
        if self.worker_thread is not None:
            self.worker_thread.deleteLater()
            self.worker_thread = None

    def on_finished(self, output_path: str) -> None:
        self.append_log("העיבוד הסתיים בהצלחה")
        self.append_log(f"קובץ הפלט: {output_path}")
        self._refresh_stats()
        QMessageBox.information(self, APP_TITLE, f"הדוח נשמר בהצלחה:\n{output_path}")

    def on_error(self, error_message: str) -> None:
        self.append_log(f"שגיאה: {error_message}")
        _logger.error(error_message)
        QMessageBox.critical(self, APP_TITLE, error_message)

    def choose_agent_table(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self, "בחר קובץ טבלת סוכנים", "", "קבצי Excel (*.xlsx *.xls)"
        )
        if not file_path:
            return
        try:
            load_agent_table(file_path)
        except Exception as exc:
            QMessageBox.critical(self, APP_TITLE, f"שגיאה בטעינת טבלת סוכנים:\n{exc}")
            return
        self.agent_table_path = file_path
        self.agent_path_edit.setText(file_path)
        save_agents_config(
            os.path.join(_APP_DIR, AGENTS_CONFIG_FILE_NAME),
            {"agent_table_path": file_path},
        )
        self.append_log(f"טבלת סוכנים נטענה: {os.path.basename(file_path)}")
        self._refresh_stats()

    def _refresh_stats(self) -> None:
        output_dir = self.output_path_edit.text().strip()
        if not output_dir or not self.agent_table_path:
            self.stat_matched.reset()
            self.stat_unmatched.reset()
            return
        try:
            agent_df = load_agent_table(self.agent_table_path)
            rows_map, unmatched_df = build_agent_rows_map(output_dir, agent_df, lambda _: None)
        except Exception:
            self.stat_matched.reset()
            self.stat_unmatched.reset()
            return
        matched_agents = len(rows_map)
        matched_records = sum(len(v) for v in rows_map.values())
        unmatched_agents = unmatched_df["_agent_key"].nunique() if not unmatched_df.empty else 0
        self.stat_matched.set_value(matched_agents, f"{matched_records} רשומות")
        self.stat_unmatched.set_value(unmatched_agents)

    def start_email_worker(self) -> None:
        output_dir = self.output_path_edit.text().strip()
        if not output_dir:
            QMessageBox.warning(self, APP_TITLE, "יש לבחור תיקיית פלט לפני הכנת המיילים")
            return
        if not self.agent_table_path:
            QMessageBox.warning(self, APP_TITLE, "יש לטעון טבלת סוכנים לפני הכנת המיילים")
            return
        if self.email_worker_thread is not None and self.email_worker_thread.isRunning():
            QMessageBox.information(self, APP_TITLE, "כבר מתבצעת הכנת מיילים, יש להמתין לסיום")
            return
        self.set_busy(True)
        self.email_worker_thread = QThread(self)
        self.email_worker = EmailWorker(output_dir, self.agent_table_path)
        self.email_worker.moveToThread(self.email_worker_thread)
        self.email_worker_thread.started.connect(self.email_worker.run)
        self.email_worker.log_message.connect(self.append_log)
        self.email_worker.finished.connect(self.on_emails_finished)
        self.email_worker.error.connect(self.on_email_error)
        self.email_worker.finished.connect(self.email_worker_thread.quit)
        self.email_worker.error.connect(self.email_worker_thread.quit)
        self.email_worker_thread.finished.connect(self.cleanup_email_thread)
        self.email_worker_thread.start()

    def cleanup_email_thread(self) -> None:
        self.set_busy(False)
        if self.email_worker is not None:
            self.email_worker.deleteLater()
            self.email_worker = None
        if self.email_worker_thread is not None:
            self.email_worker_thread.deleteLater()
            self.email_worker_thread = None

    def on_emails_finished(self, count: int) -> None:
        self.append_log(f"הכנת מיילים הסתיימה — נוצרו {count} טיוטות")
        QMessageBox.information(self, APP_TITLE, f"נוצרו {count} טיוטות מייל ב-Outlook")

    def on_email_error(self, error_message: str) -> None:
        self.append_log(f"שגיאה בהכנת מיילים: {error_message}")
        QMessageBox.critical(self, APP_TITLE, error_message)

    def start_import_worker(self) -> None:
        output_dir = self.output_path_edit.text().strip()
        if not output_dir:
            QMessageBox.warning(self, APP_TITLE, "יש לבחור תיקיית פלט לפני קליטת התגובות")
            return
        if self.import_worker_thread is not None and self.import_worker_thread.isRunning():
            QMessageBox.information(self, APP_TITLE, "כבר מתבצעת קליטת תגובות, יש להמתין לסיום")
            return
        self.set_busy(True)
        self.import_worker_thread = QThread(self)
        self.import_worker = ImportWorker(output_dir)
        self.import_worker.moveToThread(self.import_worker_thread)
        self.import_worker_thread.started.connect(self.import_worker.run)
        self.import_worker.log_message.connect(self.append_log)
        self.import_worker.finished.connect(self.on_import_finished)
        self.import_worker.error.connect(self.on_import_error)
        self.import_worker.finished.connect(self.import_worker_thread.quit)
        self.import_worker.error.connect(self.import_worker_thread.quit)
        self.import_worker_thread.finished.connect(self.cleanup_import_thread)
        self.import_worker_thread.start()

    def cleanup_import_thread(self) -> None:
        self.set_busy(False)
        if self.import_worker is not None:
            self.import_worker.deleteLater()
            self.import_worker = None
        if self.import_worker_thread is not None:
            self.import_worker_thread.deleteLater()
            self.import_worker_thread = None

    def on_import_finished(self, count: int) -> None:
        self.append_log(f"קליטת תגובות סוכנים הסתיימה — נקלטו {count} הערות")
        QMessageBox.information(self, APP_TITLE, f"נקלטו {count} הערות סוכנים בהצלחה")

    def on_import_error(self, error_message: str) -> None:
        self.append_log(f"שגיאה בקליטת תגובות: {error_message}")
        QMessageBox.critical(self, APP_TITLE, error_message)


def _handle_uncaught_exception(exc_type, exc_value, exc_tb):
    """Log unhandled exceptions to file before the process exits."""
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_tb)
        return
    _logger.critical("קריסה לא מטופלת", exc_info=(exc_type, exc_value, exc_tb))


def main() -> None:
    sys.excepthook = _handle_uncaught_exception
    _logger.info("=" * 60)
    _logger.info("הפעלת הכלי")
    app = QApplication(sys.argv)
    app.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
